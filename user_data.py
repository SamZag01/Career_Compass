from openpyxl import load_workbook
import json
import os
from trainingmodel import train_model
from flask import render_template, request

# def save_data():
#     print("Scores updated.")
#     with open('scores.json', 'r') as file:
#         scores = json.load(file)
#     with open('suggested_careers.json', 'r') as file:
#         suggested_careers = json.load(file)
#     print("Scores:", scores)
#     print("Suggested Careers:", suggested_careers)
#     return render_template('index.html', scores=scores, suggested_careers=suggested_careers)

def update_data():
    print("update_data called")
    with open('data.json', 'r') as file:
        data = json.load(file)
        print(data)
    # Split the data array into two separate lists
    text_data = data[0::2]  # Elements at even indices
    int_data = data[1::2]  # Elements at odd indices
    #write location of data/test_data
    location = "static/data/test_data.xlsx"
    if not os.path.exists(location):
        raise FileNotFoundError(f"File not found at location: {location}")
        exit(0)

    #open the workbook
    wb = load_workbook(location,read_only=False)
    sheet = wb.active

    # Write text_data to the first row
    for col_index, value in enumerate(text_data):
        sheet.cell(row=1, column=col_index + 1, value=value)

    # Write int_data to the second row
    for col_index, value in enumerate(int_data):
        sheet.cell(row=2, column=col_index + 1, value=value)

    wb.save(location)
    print(f"Data from {data} has been written to {location}")
    train_model()
    # save_data()
    print("End of line in user_data")


